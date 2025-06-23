import json
import os
import sys
from openpyxl import load_workbook

def get_output_path():
    # 获取上层文件夹的config.py中的lanota_table_full_path
    config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config.py')
    
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"找不到config.py文件: {config_path}")
    
    # 动态导入config模块
    import importlib.util
    spec = importlib.util.spec_from_file_location("config", config_path)
    config = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(config)
    
    if not hasattr(config, 'lanota_table_full_path'):
        raise AttributeError("config.py中没有定义lanota_table_full_path")
    
    return config.lanota_table_full_path

def excel_to_json(input_path, sheet_name="Sheet1"):
    # 加载工作簿和工作表
    wb = load_workbook(input_path)
    sheet = wb[sheet_name]
    
    result = {}
    
    # 遍历每一行
    for row in sheet.iter_rows(values_only=True):
        if not row or not row[0]:  # 跳过空行和空键
            continue
            
        key = str(row[0])  # A列作为主键，确保转为字符串
        data = {}
        
        # 从C列开始，每两列为一组键值对
        for i in range(2, len(row), 2):
            if i + 1 < len(row) and row[i] and row[i+1]:  # 确保有键和值
                data[str(row[i]) if row[i] else ""] = str(row[i+1]) if row[i+1] else ""
        
        if data:  # 如果有数据才添加到结果中
            result[key] = data
    
    return result

def main():
    # 获取用户输入的Excel文件路径
    input_path = input("请输入Excel文件路径: ").strip('"\' ')  # 去除可能的引号和空格
    
    if not os.path.exists(input_path):
        print(f"错误: 文件不存在 - {input_path}")
        return
    
    try:
        # 获取输出路径
        output_path = get_output_path()
        output_dir = os.path.dirname(output_path)
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        # 转换数据
        json_data = excel_to_json(input_path)
        
        # 写入JSON文件
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
            
        print(f"转换完成，结果已保存到: {output_path}")
        
    except Exception as e:
        print(f"发生错误: {str(e)}")
        if hasattr(e, '__traceback__'):
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()